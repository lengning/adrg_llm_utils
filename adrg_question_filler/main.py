#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Agent to fill in yes/no questions in ADRG template.

This script:
1. Identifies yes/no questions in the ADRG template
2. Attempts to answer them using data from pipeline files
3. Skips questions that cannot be answered automatically (keeps <Yes/No> placeholder)
4. Generates a filled template with answers
"""

import argparse
import json
import os
import re
import sys
from pathlib import Path
from typing import Dict, List, Tuple, Optional
import xml.etree.ElementTree as ET

from langchain_openai import ChatOpenAI
from langchain_core.prompts import ChatPromptTemplate
from langchain_core.output_parsers import StrOutputParser

ROOT_DIR = Path(__file__).resolve().parents[1]

# ========= Question Patterns =========
YESNO_PATTERN = re.compile(
    r'^(.+?)\n<Yes/No>(.*)$',
    re.MULTILINE | re.IGNORECASE
)

# ========= Prompts =========
QUESTION_ANSWERING_PROMPT = ChatPromptTemplate.from_messages([
    ("system",
     "You are an expert in clinical trial data analysis and ADaM dataset standards. "
     "Your task is to answer yes/no questions about clinical trial analysis datasets "
     "based on the provided data files and documentation.\n\n"
     "Given a question and relevant data, you must:\n"
     "1. Carefully analyze all provided data\n"
     "2. Determine if you have sufficient information to answer the question\n"
     "3. If you can answer, respond with 'Yes' or 'No' followed by a brief explanation\n"
     "4. If you cannot answer with confidence, respond with 'CANNOT_ANSWER' followed by "
     "an explanation of what information is missing\n\n"
     "Format your response as:\n"
     "ANSWER: Yes/No/CANNOT_ANSWER\n"
     "EXPLANATION: <your explanation>\n"
     "ADDITIONAL_TEXT: <any additional context or details to include after the Yes/No>"),
    ("human",
     "Question: {question}\n\n"
     "Available Data:\n{data_context}\n\n"
     "Please answer the question based on the data provided.")
])


# ========= LLM Builder =========
def build_llm(model="gpt-4o-mini", temperature=0):
    """Build language model instance."""
    api_key = os.environ.get("OPENAI_API_KEY")
    if api_key:
        return ChatOpenAI(model=model, temperature=temperature, api_key=api_key)
    return ChatOpenAI(model=model, temperature=temperature)


# ========= Question Extraction =========
def extract_yesno_questions(template_path: Path) -> List[Tuple[str, str, str]]:
    """
    Extract yes/no questions from the ADRG template.

    Returns:
        List of tuples (line_number, question_text, existing_text)
    """
    template_text = template_path.read_text(encoding='utf-8')
    questions = []

    # Split into lines and search for yes/no patterns
    lines = template_text.split('\n')
    i = 0
    while i < len(lines):
        line = lines[i]
        # Look for lines ending with <Yes/No>
        if '<Yes/No>' in line or '<yes/no>' in line.lower():
            # Get the question (might span multiple lines)
            question_parts = []
            # Look backwards to find the start of the question
            j = i
            while j >= 0:
                current_line = lines[j].strip()
                if current_line.startswith('-') or current_line.startswith('#'):
                    question_parts.insert(0, current_line)
                    break
                elif current_line and not current_line.startswith('('):
                    question_parts.insert(0, current_line)
                    j -= 1
                else:
                    break

            if question_parts:
                question_text = ' '.join(question_parts)
                # Remove the <Yes/No> marker
                question_text = re.sub(r'<Yes/No>.*$', '', question_text, flags=re.IGNORECASE).strip()
                # Extract any existing text after <Yes/No>
                match = re.search(r'<Yes/No>\s*(.+)?', line, re.IGNORECASE)
                existing_text = match.group(1) if match and match.group(1) else ''
                questions.append((i + 1, question_text, existing_text))
        i += 1

    return questions


# ========= Input File Readers =========
def read_xml_file(xml_path: Path) -> str:
    """Extract useful information from XML file (e.g., define.xml)."""
    try:
        tree = ET.parse(xml_path)
        root = tree.getroot()

        # Remove namespace prefixes for easier parsing
        for elem in root.iter():
            if '}' in elem.tag:
                elem.tag = elem.tag.split('}', 1)[1]

        # Extract key information as text
        info_parts = []

        # Get study metadata
        study = root.find('.//Study')
        if study is not None:
            study_name = study.find('GlobalVariables/StudyName')
            if study_name is not None:
                info_parts.append(f"Study Name: {study_name.text}")

            protocol_name = study.find('GlobalVariables/ProtocolName')
            if protocol_name is not None:
                info_parts.append(f"Protocol Name: {protocol_name.text}")

        # Get dataset information
        datasets = root.findall('.//ItemGroupDef')
        if datasets:
            info_parts.append(f"\nDatasets ({len(datasets)}):")
            for ds in datasets[:20]:  # Limit to first 20
                name = ds.get('Name', 'Unknown')
                label = ds.get('Label', '')
                info_parts.append(f"  - {name}: {label}")

        # Get variable information (sample)
        variables = root.findall('.//ItemDef')
        if variables:
            info_parts.append(f"\nVariables (sample of {min(len(variables), 50)}):")
            for var in variables[:50]:  # Limit to first 50
                name = var.get('Name', 'Unknown')
                label = var.get('Label', '')
                data_type = var.get('DataType', '')
                info_parts.append(f"  - {name} ({data_type}): {label}")

        return '\n'.join(info_parts)
    except Exception as e:
        return f"Error reading XML file: {str(e)}"


def read_xlsx_file(xlsx_path: Path) -> str:
    """Extract useful information from XLSX file (e.g., ADaM spec)."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

        info_parts = []
        info_parts.append(f"Excel file sheets: {', '.join(wb.sheetnames)}")

        # Read key sheets
        for sheet_name in ['Datasets', 'Variables', 'Methods']:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                info_parts.append(f"\n{sheet_name} sheet:")

                # Read header row
                rows = list(ws.iter_rows(values_only=True))
                if rows:
                    header = rows[0]
                    info_parts.append(f"  Columns: {', '.join(str(h) for h in header if h)}")

                    # Read sample rows
                    sample_count = min(20, len(rows) - 1)
                    if sample_count > 0:
                        info_parts.append(f"  Sample rows ({sample_count}):")
                        for row in rows[1:sample_count + 1]:
                            # Format row data
                            row_data = [str(cell) if cell is not None else '' for cell in row]
                            info_parts.append(f"    {' | '.join(row_data[:5])}")  # First 5 columns

        wb.close()
        return '\n'.join(info_parts)
    except ImportError:
        return "openpyxl not installed - cannot read XLSX file"
    except Exception as e:
        return f"Error reading XLSX file: {str(e)}"


# ========= Data Context Building =========
def build_data_context(config: Dict, base_path: Path) -> str:
    """Build context from all available data files."""
    context_parts = []

    # Helper to resolve paths
    def resolve_path(path_str: str) -> Path:
        path = Path(path_str).expanduser()
        if not path.is_absolute():
            path = base_path / path
        return path

    # Read protocol description if available
    if 'protocol_retrieve' in config:
        protocol_cfg = config['protocol_retrieve']
        out_path = resolve_path(protocol_cfg.get('out', 'outputs/protocol_description.md'))
        if out_path.exists():
            context_parts.append("=== PROTOCOL INFORMATION ===")
            context_parts.append(out_path.read_text(encoding='utf-8'))

    # Read variable descriptions
    if 'adam_info' in config:
        adam_cfg = config['adam_info']
        out_path = resolve_path(adam_cfg.get('out', 'outputs/var_descriptions.csv'))
        if out_path.exists():
            context_parts.append("\n=== VARIABLE DESCRIPTIONS ===")
            context_parts.append(out_path.read_text(encoding='utf-8'))

        deps_path = resolve_path(adam_cfg.get('deps_out', 'outputs/dataset_dependencies.csv'))
        if deps_path.exists():
            context_parts.append("\n=== DATASET DEPENDENCIES ===")
            context_parts.append(deps_path.read_text(encoding='utf-8'))

    # Read analysis outputs (R script analysis)
    if 'var_filter' in config:
        var_cfg = config['var_filter']
        out_path = resolve_path(var_cfg.get('out', 'outputs/output_var_filter_folder.csv'))
        if out_path.exists():
            context_parts.append("\n=== ANALYSIS PROGRAMS AND VARIABLES USED ===")
            context_parts.append(out_path.read_text(encoding='utf-8'))

    # Read R package info
    if 'renv_to_table' in config:
        renv_cfg = config['renv_to_table']
        out_path = resolve_path(renv_cfg.get('out', 'outputs/r_pkg_versions.csv'))
        if out_path.exists():
            context_parts.append("\n=== R PACKAGES USED ===")
            context_parts.append(out_path.read_text(encoding='utf-8'))

    # Read standards info
    if 'sdtm_medra_version' in config:
        sdtm_cfg = config['sdtm_medra_version']
        out_path = resolve_path(sdtm_cfg.get('out', 'outputs/standards_from_define.csv'))
        if out_path.exists():
            context_parts.append("\n=== STANDARDS AND VERSIONS ===")
            context_parts.append(out_path.read_text(encoding='utf-8'))

    # Read INPUT files for additional context

    # Read define.xml if specified
    if 'sdtm_medra_version' in config:
        sdtm_cfg = config['sdtm_medra_version']
        if 'define' in sdtm_cfg:
            define_path = resolve_path(sdtm_cfg['define'])
            if define_path.exists() and define_path.suffix.lower() == '.xml':
                context_parts.append("\n=== DEFINE.XML METADATA ===")
                context_parts.append(read_xml_file(define_path))

    # Read ADaM spec XLSX if specified
    if 'adam_info' in config:
        adam_cfg = config['adam_info']
        if 'spec' in adam_cfg:
            spec_path = resolve_path(adam_cfg['spec'])
            if spec_path.exists() and spec_path.suffix.lower() in ['.xlsx', '.xls']:
                context_parts.append("\n=== ADAM SPECIFICATION (XLSX) ===")
                context_parts.append(read_xlsx_file(spec_path))

    # Read renv.lock if specified
    if 'renv_to_table' in config:
        renv_cfg = config['renv_to_table']
        if 'renv' in renv_cfg:
            renv_path = resolve_path(renv_cfg['renv'])
            if renv_path.exists():
                try:
                    renv_content = renv_path.read_text(encoding='utf-8')
                    # Parse renv.lock JSON
                    renv_data = json.loads(renv_content)
                    packages = renv_data.get('Packages', {})
                    context_parts.append(f"\n=== R ENVIRONMENT (renv.lock) ===")
                    context_parts.append(f"R Version: {renv_data.get('R', {}).get('Version', 'unknown')}")
                    context_parts.append(f"Number of packages: {len(packages)}")
                    context_parts.append(f"Key packages: {', '.join(list(packages.keys())[:20])}")
                except Exception as e:
                    context_parts.append(f"\n=== R ENVIRONMENT (renv.lock) ===\nError reading: {e}")

    # Read sample R scripts if specified
    if 'var_filter' in config:
        var_cfg = config['var_filter']
        if 'folder' in var_cfg:
            folder_path = resolve_path(var_cfg['folder'])
            if folder_path.exists() and folder_path.is_dir():
                r_files = list(folder_path.glob('*.r')) + list(folder_path.glob('*.R'))
                if r_files:
                    context_parts.append(f"\n=== R ANALYSIS SCRIPTS ===")
                    context_parts.append(f"Number of R scripts: {len(r_files)}")
                    context_parts.append(f"Script names: {', '.join(f.name for f in r_files[:10])}")
                    # Read first script as sample
                    if r_files:
                        sample_script = r_files[0]
                        try:
                            script_content = sample_script.read_text(encoding='utf-8')
                            # Get first 50 lines as sample
                            lines = script_content.split('\n')[:50]
                            context_parts.append(f"\nSample script ({sample_script.name}):")
                            context_parts.append('\n'.join(lines))
                        except Exception:
                            pass
        elif 'file' in var_cfg:
            file_path = resolve_path(var_cfg['file'])
            if file_path.exists():
                try:
                    script_content = file_path.read_text(encoding='utf-8')
                    lines = script_content.split('\n')[:50]
                    context_parts.append(f"\n=== R ANALYSIS SCRIPT ({file_path.name}) ===")
                    context_parts.append('\n'.join(lines))
                except Exception:
                    pass

    return '\n\n'.join(context_parts)


# ========= Question Answering =========
def answer_question(question: str, data_context: str, llm) -> Dict[str, str]:
    """
    Attempt to answer a yes/no question using available data.

    Returns:
        Dict with keys: 'answer' (Yes/No/CANNOT_ANSWER), 'explanation', 'additional_text'
    """
    prompt = QUESTION_ANSWERING_PROMPT
    chain = prompt | llm | StrOutputParser()

    try:
        response = chain.invoke({
            "question": question,
            "data_context": data_context
        })

        # Parse response
        answer_match = re.search(r'ANSWER:\s*(Yes|No|CANNOT_ANSWER)', response, re.IGNORECASE)
        explanation_match = re.search(r'EXPLANATION:\s*(.+?)(?=ADDITIONAL_TEXT:|$)', response, re.DOTALL | re.IGNORECASE)
        additional_match = re.search(r'ADDITIONAL_TEXT:\s*(.+)$', response, re.DOTALL | re.IGNORECASE)

        answer = answer_match.group(1) if answer_match else 'CANNOT_ANSWER'
        explanation = explanation_match.group(1).strip() if explanation_match else ''
        additional_text = additional_match.group(1).strip() if additional_match else ''

        return {
            'answer': answer,
            'explanation': explanation,
            'additional_text': additional_text
        }
    except Exception as e:
        print(f"Error answering question: {e}", file=sys.stderr)
        return {
            'answer': 'CANNOT_ANSWER',
            'explanation': f'Error during processing: {str(e)}',
            'additional_text': ''
        }




# ========= Template Filling =========
def fill_template(
    template_path: Path,
    questions_and_answers: List[Tuple[int, str, str, str]]
) -> str:
    """
    Fill in the template with answers.

    Args:
        template_path: Path to template file
        questions_and_answers: List of (line_num, question, answer, additional_text)

    Returns:
        Filled template text
    """
    template_text = template_path.read_text(encoding='utf-8')
    lines = template_text.split('\n')

    # Create a mapping of line numbers to answers
    answer_map = {}
    for line_num, question, answer, additional_text in questions_and_answers:
        answer_map[line_num - 1] = (answer, additional_text)  # Convert to 0-indexed

    # Replace yes/no placeholders
    for i, line in enumerate(lines):
        if i in answer_map:
            answer, additional_text = answer_map[i]
            # Replace <Yes/No> with the actual answer
            # Ensure answer is clearly visible with proper formatting
            if answer == '<Yes/No>':
                # Keep placeholder as-is if question couldn't be answered
                replacement = answer
            else:
                # Format: "Yes." or "No." followed by additional text
                replacement = f"**{answer}.**"
                if additional_text:
                    replacement += f" {additional_text}"
            lines[i] = re.sub(r'<Yes/No>.*$', replacement, line, flags=re.IGNORECASE)

    return '\n'.join(lines)


# ========= Main Function =========
def main():
    ap = argparse.ArgumentParser(
        description="Fill yes/no questions in ADRG template using pipeline data."
    )
    ap.add_argument(
        "--config",
        type=Path,
        default=ROOT_DIR / "adrg_doc" / "example_pipeline_config.json",
        help="Path to pipeline configuration JSON file"
    )
    ap.add_argument(
        "--template",
        type=Path,
        default=ROOT_DIR / "adrg_doc" / "adrg-template.qmd",
        help="Path to ADRG template file"
    )
    ap.add_argument(
        "--out",
        type=Path,
        default=ROOT_DIR / "outputs" / "adrg-filled.qmd",
        help="Output file path for filled template"
    )
    ap.add_argument(
        "--model",
        default="gpt-4o-mini",
        help="LLM model to use for question answering"
    )
    args = ap.parse_args()

    # Validate inputs
    if not args.config.exists():
        print(f"Error: Config file not found: {args.config}", file=sys.stderr)
        sys.exit(1)
    if not args.template.exists():
        print(f"Error: Template file not found: {args.template}", file=sys.stderr)
        sys.exit(1)

    # Load configuration
    with args.config.open('r', encoding='utf-8') as f:
        config = json.load(f)

    # Extract questions from template
    print("Extracting yes/no questions from template...", file=sys.stderr)
    questions = extract_yesno_questions(args.template)
    print(f"Found {len(questions)} yes/no questions", file=sys.stderr)

    # Build data context
    print("Building data context from pipeline files...", file=sys.stderr)
    data_context = build_data_context(config, ROOT_DIR)

    if not data_context.strip():
        print("Warning: No data files found. Answers will need to be provided manually.", file=sys.stderr)

    # Build LLM
    llm = build_llm(model=args.model)

    # Process each question
    questions_and_answers = []
    for line_num, question, existing_text in questions:
        print(f"\nProcessing question at line {line_num}:", file=sys.stderr)
        print(f"  {question}", file=sys.stderr)

        # Try to answer automatically
        result = answer_question(question, data_context, llm)

        if result['answer'] == 'CANNOT_ANSWER':
            print(f"  Skipping (cannot answer automatically)", file=sys.stderr)
            # Keep the original <Yes/No> placeholder
            questions_and_answers.append((line_num, question, '<Yes/No>', existing_text))
        else:
            print(f"  Answer: {result['answer']}", file=sys.stderr)
            print(f"  Explanation: {result['explanation']}", file=sys.stderr)
            questions_and_answers.append((
                line_num,
                question,
                result['answer'],
                result['additional_text'] or existing_text
            ))

    # Fill template
    print("\nFilling template with answers...", file=sys.stderr)
    filled_text = fill_template(args.template, questions_and_answers)

    # Write output
    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(filled_text, encoding='utf-8')
    print(f"\nFilled template written to: {args.out}", file=sys.stderr)


if __name__ == "__main__":
    main()
